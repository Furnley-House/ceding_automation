from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill

FP = r"C:\Users\RevathyS\Downloads\Ceding_Automation_UAT_Test_Tracker_v2.xlsx"

REMOVE_IDS = {
    "TC-IMP-03",  # provider auto-create
    "TC-S4-03",   # CA approve a field (only PP)
    "TC-S4-05",   # CA request review (only PP)
    "TC-S5-05",   # transcript source hover (not in UI)
    "TC-S8-03",   # standalone comment (only review-with-comment exists)
    "TC-S9-01",   # export preview (no preview in UI)
    "TC-SP-01",   # On Hold (not available)
    "TC-SP-02",   # Resume from On Hold (not available)
}

wb = load_workbook(FP)
ws = wb["Test Script"]

START_ROW = 5
# find rows to delete
rows_to_delete = []
for r in range(START_ROW, ws.max_row + 1):
    tid = ws.cell(row=r, column=1).value
    if tid in REMOVE_IDS:
        rows_to_delete.append(r)

# delete bottom-up so indices stay valid
for r in reversed(rows_to_delete):
    ws.delete_rows(r, 1)

removed = len(rows_to_delete)
new_last_row = ws.max_row
new_total = new_last_row - START_ROW + 1
print(f"Removed {removed} rows. New last row: {new_last_row}. New total: {new_total}")

# Re-stripe section banding (rows shifted)
prev_area = None
toggle = False
stripe_a = PatternFill("solid", start_color="EEF1F8")
stripe_b = PatternFill("solid", start_color="FFFFFF")
for r in range(START_ROW, new_last_row + 1):
    area = ws.cell(row=r, column=2).value
    if area != prev_area:
        toggle = not toggle
        prev_area = area
    fill = stripe_a if toggle else stripe_b
    for c in range(1, 8):
        ws.cell(row=r, column=c).fill = fill

# ---------- update Summary ----------
sm = wb["Summary"]
sm["B5"] = new_total
rng = f"'Test Script'!N{START_ROW}:N{new_last_row}"
sm["B6"] = f'=COUNTIF({rng},"PASS")'
sm["B7"] = f'=COUNTIF({rng},"FAIL")'
sm["B8"] = f'=COUNTIF({rng},"BLOCKED")'
sm["B9"] = f'=COUNTIF({rng},"REVIEW")'
sm["B10"] = f"=B5-(B6+B7+B8+B9)"
sm["B11"] = "=IFERROR((B6+B7+B8+B9)/B5,0)"
sm["B12"] = "=IFERROR(B6/(B6+B7+B8+B9),0)"

for i, col in enumerate([("H","I"), ("J","K"), ("L","M")]):
    tester_row = 16 + i
    result_col = col[0]
    rng_t = f"'Test Script'!{result_col}{START_ROW}:{result_col}{new_last_row}"
    sm.cell(row=tester_row, column=2, value=f'=COUNTIF({rng_t},"Pass")')
    sm.cell(row=tester_row, column=3, value=f'=COUNTIF({rng_t},"Fail")')
    sm.cell(row=tester_row, column=4, value=f'=COUNTIF({rng_t},"Blocked")')
    sm.cell(row=tester_row, column=5, value=f'=COUNTA({rng_t})')

wb.save(FP)
print("Saved:", FP)
