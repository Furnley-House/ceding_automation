"""Generate Zoho Sprints bulk-import workbook from sprint-plan-draft.md."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = r"C:\Users\RevathyS\Documents\ceding_automation\ceding_automation\ceding_automation\ceding_automation\project-context\zoho-sprints-import.xlsx"

# ── Styles ────────────────────────────────────────────────
NAVY = "0D1B2A"
TEAL = "00C2CB"
LIGHT_GREY = "F4F4F4"
WHITE = "FFFFFF"

HEADER_FONT = Font(name="Arial", size=11, bold=True, color=WHITE)
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
SUBHEADER_FONT = Font(name="Arial", size=11, bold=True, color=NAVY)
SUBHEADER_FILL = PatternFill("solid", fgColor=LIGHT_GREY)
DATA_FONT = Font(name="Arial", size=10)
TITLE_FONT = Font(name="Arial", size=14, bold=True, color=NAVY)
NOTE_FONT = Font(name="Arial", size=10, italic=True, color="666666")

THIN = Side(border_style="thin", color="CCCCCC")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WRAP_TOP = Alignment(wrap_text=True, vertical="top", horizontal="left")
WRAP_CTR = Alignment(wrap_text=True, vertical="center", horizontal="center")


def header(ws, row, labels, widths=None):
    for c, label in enumerate(labels, 1):
        cell = ws.cell(row=row, column=c, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP_CTR
        cell.border = BOX
    if widths:
        for c, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[row].height = 28
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def row(ws, r, values):
    for c, v in enumerate(values, 1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = DATA_FONT
        cell.alignment = WRAP_TOP
        cell.border = BOX


def title(ws, r, text, width=10):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=width)
    cell = ws.cell(row=r, column=1, value=text)
    cell.font = TITLE_FONT
    cell.alignment = Alignment(vertical="center", horizontal="left")


def note(ws, r, text, width=10):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=width)
    cell = ws.cell(row=r, column=1, value=text)
    cell.font = NOTE_FONT
    cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")


# ── Owner mapping (display name → likely email; email column left blank where unknown) ──
OWNERS = {
    "Revathy S": "revathy.s@furnleyhouse.co.uk",
    "Revathy": "revathy.s@furnleyhouse.co.uk",
    "Nishant R": "nishant@furnleyhouse.co.uk",
    "Nishant": "nishant@furnleyhouse.co.uk",
    "Srinath K": "srinath@furnleyhouse.co.uk",
    "Srinath": "srinath@furnleyhouse.co.uk",
    "Daniel Worthing": "",
    "Chris Vaughan": "",
    "All": "",
}

# ── Workbook ──────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)

# ===========================================================================
# Sheet 1 — README
# ===========================================================================
ws = wb.create_sheet("README")
ws.sheet_view.showGridLines = False
title(ws, 1, "Ceding Automation — Zoho Sprints Bulk-Import Workbook")
note(ws, 2, "Generated from project-context/sprint-plan-draft.md on 2026-05-07.")

ws.cell(row=4, column=1, value="How to use this workbook").font = SUBHEADER_FONT
instructions = [
    "1. Create a new project in Zoho Sprints called 'Ceding Automation' (one-time, via the Zoho Sprints web UI).",
    "2. Create 7 sprints inside that project — names and dates are listed on sheet 'Project Setup'. Sprint 1 starts 7 May 2026; Sprint 7 ends 1 July 2026.",
    "3. (Optional) Create epics from sheet 'Epics' before importing items, so items can be linked to epics on import. If you skip this step, import items first and link epics afterwards via the Sprints UI.",
    "4. For each sprint, open the sprint in Sprints UI, click 'Backlog' → 'Import' → select the matching 'Sprint N Items' sheet (saved as CSV or XLSX). Map columns as prompted.",
    "5. Apply tags (TR-XX / FR-XX) per row after import — Sprints' bulk import doesn't always carry tags through; the 'Tags' column in each sprint sheet lists what to apply.",
    "6. Set inter-task dependencies via the Sprints UI using the 'Depends On' column as guidance.",
]
for i, line in enumerate(instructions, start=5):
    ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=10)
    cell = ws.cell(row=i, column=1, value=line)
    cell.font = DATA_FONT
    cell.alignment = WRAP_TOP
    ws.row_dimensions[i].height = 30

ws.cell(row=12, column=1, value="Hard gates / blockers (do not push past these)").font = SUBHEADER_FONT
gates = [
    "TR-09 GDPR sign-off (Owner: John / Lee). HARD BLOCKER for Sprint 6 cutover. No live client data without this.",
    "NFR-04 AI accuracy: ≥85% HIGH-confidence on 50+ PDFs from 6+ providers. HARD GATE between Sprint 4 and Sprint 5.",
    "Provider Directory data load (Sprint 3, B12) depends on Aruna delivering verified file (Q-06).",
    "Policy reference prefix patterns (Sprint 3, B12) depend on Aruna confirming patterns (Q-07).",
]
for i, line in enumerate(gates, start=13):
    ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=10)
    cell = ws.cell(row=i, column=1, value=line)
    cell.font = DATA_FONT
    cell.alignment = WRAP_TOP
    ws.row_dimensions[i].height = 30

ws.cell(row=18, column=1, value="Sheet index").font = SUBHEADER_FONT
sheets_idx = [
    ("Project Setup", "Project name + 7 sprint definitions (name, goal, dates, branch)."),
    ("Epics", "Epic-level groupings — one per TR cluster per sprint. Optional pre-create."),
    ("Sprint 1 Items", "Foundation & Backlog Capture (7–14 May) — 30+ items including completed history."),
    ("Sprint 2 Items", "Production Integration Wiring (15–22 May) — 19 items."),
    ("Sprint 3 Items", "Azure Deploy & Pre-Launch Polish (23 May – 1 June) — 26 items."),
    ("Sprint 4 Items", "Internal Testing Chennai (1–7 June) — 7 items."),
    ("Sprint 5 Items", "Business UAT (8–15 June) — 7 items."),
    ("Sprint 6 Items", "Production Launch & Hypercare (16–27 June) — 10 items."),
    ("Sprint 7 Items", "Success Review & Closeout (28 June – 1 July) — 6 items."),
    ("Branch Plan", "B1–B20 git branches mapped to sprints."),
]
header(ws, 19, ["Sheet", "Contents"], widths=[24, 80])
for i, (name, desc) in enumerate(sheets_idx, start=20):
    row(ws, i, [name, desc])

ws.column_dimensions["A"].width = 28
for col in "BCDEFGHIJ":
    ws.column_dimensions[col].width = 12

# ===========================================================================
# Sheet 2 — Project Setup
# ===========================================================================
ws = wb.create_sheet("Project Setup")
ws.sheet_view.showGridLines = False
title(ws, 1, "Project: Ceding Automation")
note(ws, 2, "Create this project in Zoho Sprints, then create the 7 sprints below.")

header(
    ws,
    4,
    ["#", "Sprint Name", "Goal", "Start Date", "End Date", "Days", "Branch"],
    widths=[5, 38, 70, 14, 14, 8, 26],
)

sprints = [
    (1, "Sprint 1: Foundation & Backlog Capture",
     "Record completed prototype work; harden existing features; kick off Azure tenant prep.",
     "07/05/2026", "14/05/2026", 8, "feature/foundation-hardening (B1)"),
    (2, "Sprint 2: Production Integration Wiring",
     "Replace prototype stubs with production integrations (Azure OpenAI, Zoho webhook, WorkDrive, RingCentral).",
     "15/05/2026", "22/05/2026", 8, "B2–B8 (multiple)"),
    (3, "Sprint 3: Azure Deploy & Pre-Launch Polish",
     "Stand up Azure prod, run perf + load tests, deploy, code freeze on 1 June.",
     "23/05/2026", "01/06/2026", 10, "B9–B14, release/v1.0-rc cut"),
    (4, "Sprint 4: Internal Testing (Chennai)",
     "Hit the 85% AI accuracy gate (NFR-04) and clear functional testing in Chennai.",
     "01/06/2026", "07/06/2026", 7, "release/v1.0-rc + hotfix/*"),
    (5, "Sprint 5: Business UAT",
     "Five business users sign off; final pen-test; cutover plan ready.",
     "08/06/2026", "15/06/2026", 8, "release/v1.0-rc2 + hotfix/*"),
    (6, "Sprint 6: Production Launch & 10-day Hypercare",
     "Cutover 16/17 June; 10 days of hypercare monitoring.",
     "16/06/2026", "27/06/2026", 12, "release/v1.0 + hotfix/hypercare-*"),
    (7, "Sprint 7: Success Review & Closeout",
     "KPI metrics report + success meet 1 July + Phase 2 backlog grooming.",
     "28/06/2026", "01/07/2026", 4, "main only (no new dev)"),
]
for i, s in enumerate(sprints, start=5):
    row(ws, i, list(s))
    ws.row_dimensions[i].height = 38

# ===========================================================================
# Sheet 3 — Epics
# ===========================================================================
ws = wb.create_sheet("Epics")
ws.sheet_view.showGridLines = False
title(ws, 1, "Epics — one per TR/Area, owned by sprint where work happens")

header(
    ws,
    3,
    ["Epic Name", "Sprint", "TR Refs", "Description", "Owner Role"],
    widths=[44, 12, 18, 70, 22],
)

epics = [
    ("E01: Architecture & scaffold (TR-01)", 1, "TR-01",
     "Foundational app scaffold: Express + Prisma backend, React + Vite frontend, Postgres DB.",
     "backendDev / frontendDev"),
    ("E02: Project definition & KPIs (BR-05/06)", 1, "BR-05, BR-06",
     "PMO scope, KPIs, deliverables, requirements register, sign-off.",
     "Revathy S"),
    ("E03: Origo integration feasibility (TR-12)", 1, "TR-12",
     "Investigate Origo API; conclude URL-only in MVP.",
     "Revathy S"),
    ("E04: Lovable design (UI/UX)", 1, "UI/UX",
     "Prototype design realised in frontend.",
     "Revathy S"),
    ("E05: SSO + RBAC (TR-02/07)", 1, "TR-02, TR-07",
     "Microsoft Azure SSO with auto-provisioning; RBAC across CA/Adviser/PP/Admin.",
     "backendDev / securityDev"),
    ("E06: Case lifecycle + audit (FR-01..05)", 1, "FR-01..05",
     "10-stage workflow, status, LOA, paraplanner assign, audit trail.",
     "backendDev / frontendDev"),
    ("E07: Document upload + checklist edit (FR-06/08/13)", 1, "FR-06, FR-08, FR-13",
     "Multi-file upload, source citations, manual edit with override flag.",
     "backendDev / frontendDev"),
    ("E08: Excel + WorkDrive export (FR-17)", 1, "FR-17",
     "Excel export (3 tabs); WorkDrive upload (stub then live in S2).",
     "backendDev / frontendDev"),
    ("E09: Admin Panel (Section 13)", 1, "Section 13",
     "User Management, Provider Management, Checklist Templates panels.",
     "frontendDev / backendDev"),
    ("E10: Sprint 1 hardening", 1, "TR-07, NFR-02",
     "Zod audit, indexes, test scaffold, dev workflow docs, Azure subscription provisioning.",
     "backendDev / databaseDev / devOpsDev"),
    ("E11: AI architecture (TR-11)", 1, "TR-11, NFR-04",
     "Azure OpenAI deployment + AI training (long-running across S1–S3).",
     "Nishant R"),
    ("E12: Auto-create case via Zoho blueprint (TR-04)", 2, "TR-04, FR-03",
     "Zoho CRM blueprint webhook → auto-create case.",
     "backendDev"),
    ("E13: WorkDrive export (TR-06)", 2, "TR-06, FR-17",
     "Live WorkDrive API wiring (Client / Deal WIP folder).",
     "backendDev"),
    ("E14: RingCentral + Palindrome (TR-05)", 2, "TR-05",
     "Recording retrieval + transcription pipeline.",
     "backendDev"),
    ("E15: Origo & LOA (TR-12 / FR-21/23)", 2, "TR-12, FR-21, FR-23",
     "Stage 1 Origo URL routing; LOA template + email subjects.",
     "backendDev / frontendDev"),
    ("E16: Stage 9 paraplanner notification (FR-16)", 2, "FR-16",
     "Zoho task + in-app notification on Ready for Review.",
     "backendDev"),
    ("E17: Secrets management (TR-08)", 2, "TR-08",
     "Azure Key Vault wiring + rotation policy.",
     "securityDev / devOpsDev"),
    ("E18: Backend AI swap (TR-11)", 2, "TR-11",
     "Backend swap from Anthropic → Azure OpenAI.",
     "backendDev"),
    ("E19: AI prompts & accuracy (TR-11 / NFR-04)", 2, "TR-11, NFR-04",
     "Pension/ISA/GIA prompts, confidence scoring, conflict detection, call script + transcript prompts.",
     "Nishant R"),
    ("E20: Azure infrastructure (TR-01/08)", 3, "TR-01, TR-08",
     "App Service, Static Web Apps, Postgres, Blob, Insights, TLS, scaling, backups.",
     "devOpsDev / databaseDev"),
    ("E21: CI/CD pipeline (TR-01)", 3, "TR-01",
     "GitHub Actions + dev/staging/prod environments.",
     "devOpsDev"),
    ("E22: Data retention (TR-10)", 3, "TR-10",
     "1-year automated deletion job.",
     "backendDev / devOpsDev"),
    ("E23: Provider directory data (FR-19)", 3, "FR-19, Q-06, Q-07",
     "Production migration + seed; provider data load; prefix routing rules.",
     "backendDev / databaseDev"),
    ("E24: UI launch polish", 3, "UI/UX",
     "Branding, empty states, error toast, accessibility audit.",
     "frontendDev / testerDev"),
    ("E25: Security & GDPR (TR-08/09)", 3, "TR-08, TR-09",
     "Security review; GDPR sign-off pack handoff.",
     "securityDev"),
    ("E26: AI accuracy gate (NFR-04)", 3, "NFR-04",
     "Benchmark across 50 PDFs / 6+ providers; reach 85% HIGH confidence.",
     "Nishant R / testerDev"),
    ("E27: Performance & load (NFR-01/02/06)", 3, "NFR-01, NFR-02, NFR-06",
     "E2E tests, perf benchmarks, load tests for 20 concurrent users.",
     "testerDev"),
    ("E28: Cutover prep & code freeze", 3, "TR-01",
     "Deployment runbook, IT test, RC tag, handover.",
     "devOpsDev / All"),
    ("E29: AI accuracy gate (final)", 4, "NFR-04",
     "Final benchmark; sign-off from Lee + Aruna.",
     "Nishant R / testerDev"),
    ("E30: CA team functional UAT", 4, "NFR-05",
     "CA functional test plan, test data, daily triage, daily report.",
     "testerDev / All"),
    ("E31: Long-running IT bug fix", 4, "—",
     "Spans Sprints 3–6 per existing Zoho task.",
     "Nishant R / Srinath K / Revathy S"),
    ("E32: Business UAT (NFR-05)", 5, "NFR-05",
     "5-user UAT, triage, sign-off package to Lee.",
     "testerDev / All"),
    ("E33: Final security (TR-08)", 5, "TR-08",
     "Final pen-test on staging.",
     "securityDev"),
    ("E34: Cutover preparation (TR-01)", 5, "TR-01",
     "Rollback drill, final data migration, cutover script.",
     "devOpsDev / databaseDev"),
    ("E35: Cutover (TR-01)", 6, "TR-01",
     "Production cutover 16/17 Jun; smoke tests; security audit.",
     "devOpsDev / All"),
    ("E36: Hypercare (NFR-03)", 6, "NFR-03",
     "10 working days of hypercare with on-call rota and daily standup.",
     "All"),
    ("E37: KPI dashboard (BR-06)", 6, "BR-06",
     "Live SLT dashboard for cases/week, AI accuracy, calls/case.",
     "backendDev"),
    ("E38: Success metrics & closeout (BR-05/06)", 7, "BR-05, BR-06",
     "KPI extraction, retro doc, success meet 1 July.",
     "All"),
    ("E39: Phase 2 grooming", 7, "—",
     "Bond, FS/DB, Protection, EIS/VCT, SR auto.",
     "backendDev / All"),
]
for i, e in enumerate(epics, start=4):
    row(ws, i, list(e))
    ws.row_dimensions[i].height = 36

# ===========================================================================
# Sprint item sheets
# ===========================================================================
SPRINT_COLUMNS = ["#", "Item Name", "Item Type", "Priority", "Status", "Owner",
                  "Owner Email", "Start Date", "End Date", "Tags",
                  "Epic", "Branch", "Depends On", "Description"]
SPRINT_WIDTHS = [5, 50, 10, 10, 14, 22, 32, 14, 14, 22, 28, 32, 24, 70]


def add_sprint_sheet(name, items):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    title(ws, 1, name)
    header(ws, 3, SPRINT_COLUMNS, widths=SPRINT_WIDTHS)
    for i, it in enumerate(items, start=4):
        owner = it.get("owner", "")
        owner_email = OWNERS.get(owner.split(",")[0].strip(), "") if owner else ""
        row(ws, i, [
            it.get("num", i - 3),
            it["name"],
            it.get("type", "Task"),
            it.get("priority", "Medium"),
            it.get("status", "Open"),
            owner,
            owner_email,
            it.get("start", ""),
            it.get("end", ""),
            it.get("tags", ""),
            it.get("epic", ""),
            it.get("branch", ""),
            it.get("deps", ""),
            it.get("desc", ""),
        ])
        ws.row_dimensions[i].height = 56


# ───────────── Sprint 1 (7–14 May) ─────────────
S1 = [
    # PMO history (Closed)
    {"name": "Define project deliverables", "owner": "Revathy S", "priority": "High", "status": "Closed",
     "start": "18/03/2026", "end": "27/03/2026", "tags": "BR-05",
     "epic": "E02: Project definition & KPIs", "desc": "Closed (Zoho Projects history)."},
    {"name": "Define project scope (IN/OUT)", "owner": "Revathy S", "priority": "High", "status": "Closed",
     "start": "18/03/2026", "end": "27/03/2026", "tags": "BR-05",
     "epic": "E02: Project definition & KPIs", "desc": "Closed (Zoho Projects history)."},
    {"name": "Define project KPIs", "owner": "Revathy S", "priority": "High", "status": "Closed",
     "start": "27/03/2026", "end": "31/03/2026", "tags": "BR-05, BR-06",
     "epic": "E02: Project definition & KPIs", "desc": "Closed (Zoho Projects history). KPIs ratified."},
    {"name": "Explore Origo integration with API functionality", "owner": "Revathy S", "priority": "High", "status": "Closed",
     "start": "24/03/2026", "end": "31/03/2026", "tags": "TR-12",
     "epic": "E03: Origo integration feasibility", "desc": "Closed. Outcome: No API. URL link only in MVP (Q-22)."},
    {"name": "Capture project requirements (BR / FR / TR)", "owner": "Revathy S", "priority": "High", "status": "Closed",
     "start": "19/03/2026", "end": "13/04/2026", "tags": "All",
     "epic": "E02: Project definition & KPIs", "desc": "Closed. v5 sign-off doc produced."},
    {"name": "Finalise + approve requirements", "owner": "Revathy S, Daniel Worthing", "priority": "High", "status": "Closed",
     "start": "14/04/2026", "end": "16/04/2026", "tags": "All",
     "epic": "E02: Project definition & KPIs", "desc": "Closed. Requirements signed off by Lee + Daniel."},
    {"name": "Lovable design", "owner": "Revathy S", "priority": "High", "status": "Closed",
     "start": "20/03/2026", "end": "16/04/2026", "tags": "UI/UX",
     "epic": "E04: Lovable design", "desc": "Closed. Prototype design implemented in frontend."},
    {"name": "Front/Back end architecture", "owner": "Nishant R", "priority": "High", "status": "Closed",
     "start": "06/04/2026", "end": "16/04/2026", "tags": "TR-01",
     "epic": "E01: Architecture & scaffold", "desc": "Closed. Architecture document produced; stack ratified."},
    {"name": "DB set up", "owner": "Nishant R, Srinath K", "priority": "High", "status": "Closed",
     "start": "27/04/2026", "end": "07/05/2026", "tags": "TR-01",
     "epic": "E01: Architecture & scaffold", "desc": "Closed. Postgres dev DB provisioned; Prisma schema applied."},
    # Codebase work (mostly already merged)
    {"name": "Express + Prisma backend scaffold (12 routes)", "owner": "backendDev", "priority": "High", "status": "Closed",
     "tags": "TR-01", "epic": "E01: Architecture & scaffold", "branch": "main",
     "desc": "12 route modules registered (auth, cases, checklist, checklist-templates, crm, documents, fund-lines, providers, users, audit, notifications, calls). Prisma schema + migrations applied."},
    {"name": "React + Vite frontend scaffold (20 pages)", "owner": "frontendDev", "priority": "High", "status": "Closed",
     "tags": "UI/UX", "epic": "E01: Architecture & scaffold", "branch": "main",
     "desc": "20 pages routed via React Router; shadcn/ui + Tailwind theme; auth-gated routes via RoleGuard."},
    {"name": "Microsoft Azure SSO + auto-provisioning", "owner": "backendDev", "priority": "High", "status": "Closed",
     "tags": "TR-02, TR-07", "epic": "E05: SSO + RBAC", "branch": "main",
     "desc": "/auth/azure → callback → JWT issued. Unknown email auto-provisioned as CA_TEAM ACTIVE with ssoId set. Inactive users blocked. Auth race-condition fixed."},
    {"name": "RBAC across 4 roles", "owner": "backendDev", "priority": "High", "status": "Closed",
     "tags": "TR-07", "epic": "E05: SSO + RBAC", "branch": "main",
     "desc": "requireAuth + requireRole middleware on protected routes. Frontend RoleGuard mirrors. Self-protection: admins cannot demote/deactivate themselves."},
    {"name": "Case CRUD + status + LOA + chase + paraplanner assign", "owner": "backendDev", "priority": "High", "status": "Closed",
     "tags": "FR-01, FR-02, FR-25", "epic": "E06: Case lifecycle + audit", "branch": "main",
     "desc": "10-stage workflow visible per case; sequential gating; audit trail row per state change; LOA gates Stage 2."},
    {"name": "Zoho-to-DB sync on case load", "owner": "backendDev", "priority": "High", "status": "Closed",
     "tags": "TR-03, FR-02", "epic": "E06: Case lifecycle + audit", "branch": "main",
     "desc": "POST /cases/:id/sync-from-zoho re-pulls task and updates clientName, policyRef, planType, providerId, assignedToId, zohoDeepLink, zohoCaseId, clientZohoId. Provider resolved from Zoho Provider_group field. Owner email resolved → app user; auto-creates user if not in DB; never falls back to old assignee."},
    {"name": "Document upload (PDF/Word/Excel/text) multi-file", "owner": "frontendDev", "priority": "High", "status": "Closed",
     "tags": "FR-06", "epic": "E07: Document upload + checklist edit", "branch": "main",
     "desc": "Drag-and-drop multi-file upload. Each upload triggers AI extraction."},
    {"name": "Side-by-side PDF viewer + source citations", "owner": "frontendDev", "priority": "High", "status": "Closed",
     "tags": "FR-08", "epic": "E07: Document upload + checklist edit", "branch": "main",
     "desc": "Document viewer alongside checklist. Each field source citation clickable, scrolls PDF panel to exact page."},
    {"name": "Manual edit + per-field approve/review/comment", "owner": "frontendDev", "priority": "High", "status": "Closed",
     "tags": "FR-13, FR-15", "epic": "E07: Document upload + checklist edit", "branch": "main",
     "desc": "Per-field Approve / Request Review / Add Comment / Edit Value. All edits logged to audit trail."},
    {"name": "Excel export (Summary + Checklist + Audit Trail tabs)", "owner": "backendDev", "priority": "High", "status": "Closed",
     "tags": "FR-17", "epic": "E08: Excel + WorkDrive export", "branch": "main",
     "desc": "3-sheet workbook. File naming [Client]_[Provider]_[PlanRef]_[YYYYMMDD].xlsx."},
    {"name": "WorkDrive upload (stub)", "owner": "backendDev", "priority": "High", "status": "Closed",
     "tags": "FR-17", "epic": "E08: Excel + WorkDrive export", "branch": "main",
     "desc": "Stub only — production wiring in Sprint 2 (B4)."},
    {"name": "Admin Panel: User / Provider / Checklist Templates", "owner": "frontendDev", "priority": "High", "status": "Closed",
     "tags": "Section 13", "epic": "E09: Admin Panel", "branch": "main",
     "desc": "Users tab: search/filter, role inline edit, status toggle, add user dialog, self-protection. Providers tab: full CRUD, soft-delete, prefix routing. Templates tab: per plan-type, drag-reorder, type editor, dropdown options manager."},
    {"name": "Remove Supabase, switch to pure Express API", "owner": "backendDev", "priority": "High", "status": "Closed",
     "tags": "TR-01", "epic": "E10: Sprint 1 hardening", "branch": "main",
     "desc": "All Supabase references removed. Auth race-condition fix shipped (post-SSO refresh)."},
    # Currently in-progress (existing Zoho tasks)
    {"name": "Back end set up", "owner": "Srinath K, Nishant R", "priority": "High", "status": "In Progress",
     "start": "27/04/2026", "end": "07/05/2026", "tags": "TR-01",
     "epic": "E01: Architecture & scaffold",
     "desc": "20% complete (ON TRACK). All 12 route modules wired; auth middleware in place; Prisma client generated. Finishes today."},
    {"name": "AI set up", "owner": "Nishant R", "priority": "High", "status": "In Progress",
     "start": "20/04/2026", "end": "08/05/2026", "tags": "TR-11",
     "epic": "E11: AI architecture",
     "desc": "20% complete (AT RISK). Azure OpenAI deployment provisioned in Furnley tenant; endpoint accessible from backend; API key in Key Vault. Zoho task end date: 8 May."},
    {"name": "AI training (long-running, spans S1–S3)", "owner": "Nishant R, Chris Vaughan", "priority": "High", "status": "In Progress",
     "start": "28/04/2026", "end": "26/05/2026", "tags": "TR-11, NFR-04",
     "epic": "E11: AI architecture",
     "desc": "10% complete (ON TRACK). Iterative prompt tuning against benchmark PDFs; targets NFR-04 gate (≥85% HIGH on 50+ PDFs / 6+ providers) by Sprint 3 end."},
    {"name": "Integration Development (long-running, spans S1–S2)", "owner": "Srinath K, Revathy S, Nishant R", "priority": "High", "status": "Open",
     "start": "07/05/2026", "end": "21/05/2026", "tags": "TR-04, TR-05, TR-06, TR-11",
     "epic": "E11: AI architecture",
     "desc": "0% (ON TRACK). Backend ↔ AI service; Backend ↔ Zoho CRM (incl. webhook); Backend ↔ RingCentral / Palindrome; Backend ↔ WorkDrive."},
    # Sprint 1 hardening (pending)
    {"name": "Zod validation audit across all backend routes", "owner": "backendDev", "priority": "High", "status": "Open",
     "start": "08/05/2026", "end": "14/05/2026", "tags": "TR-07",
     "epic": "E10: Sprint 1 hardening", "branch": "feature/foundation-hardening (B1)",
     "desc": "AC: every POST/PUT/PATCH route has explicit Zod schema; unknown fields silently dropped; 400 returned on invalid input."},
    {"name": "Prisma indexes on frequently filtered columns", "owner": "databaseDev", "priority": "Medium", "status": "Open",
     "start": "08/05/2026", "end": "14/05/2026", "tags": "NFR-02",
     "epic": "E10: Sprint 1 hardening", "branch": "feature/foundation-hardening (B1)",
     "desc": "AC: indexes on cases.assignedToId, cases.status, cases.zohoTaskId, audit_logs.caseId, audit_logs.createdAt; migration checked in; query benchmarks improved."},
    {"name": "Test scaffold (Vitest + Supertest + Playwright) + CI gate", "owner": "testerDev", "priority": "High", "status": "Open",
     "start": "08/05/2026", "end": "14/05/2026", "tags": "NFR-01, NFR-02",
     "epic": "E10: Sprint 1 hardening", "branch": "feature/foundation-hardening (B1)",
     "desc": "AC: backend Vitest+Supertest configured; frontend Vitest+RTL configured; Playwright smoke; GHA lint+unit blocks merge on red."},
    {"name": "Local dev workflow + env-template documented", "owner": "devOpsDev", "priority": "Medium", "status": "Open",
     "start": "08/05/2026", "end": "14/05/2026", "tags": "TR-01",
     "epic": "E10: Sprint 1 hardening", "branch": "feature/foundation-hardening (B1)",
     "desc": "AC: README updated; .env.example for FE+BE; optional docker-compose.yml for local Postgres."},
    {"name": "Azure subscription / resource groups provisioned", "owner": "devOpsDev", "priority": "High", "status": "Open",
     "start": "08/05/2026", "end": "14/05/2026", "tags": "TR-01",
     "epic": "E10: Sprint 1 hardening",
     "desc": "AC: Furnley House Azure subscription confirmed; rg-ceding-dev/staging/prod created; IAM scoped (Contributor for dev team, Reader for SLT)."},
]
for n, t in enumerate(S1, start=1):
    t["num"] = n
add_sprint_sheet("Sprint 1 Items", S1)

# ───────────── Sprint 2 (15–22 May) ─────────────
S2 = [
    {"name": "Zoho CRM blueprint webhook → auto-create case", "owner": "backendDev", "priority": "High", "status": "Open",
     "start": "15/05/2026", "end": "22/05/2026", "tags": "TR-04, FR-03",
     "epic": "E12: Auto-create case via Zoho blueprint", "branch": "feature/zoho-blueprint-webhook (B3)",
     "desc": "AC: webhook authenticates Zoho callbacks; triggers on 'Request ceding' stage; idempotent on duplicate signal; audit log entry written."},
    {"name": "WorkDrive API live wiring (folder Client / Deal WIP)", "owner": "backendDev", "priority": "High", "status": "Open",
     "start": "15/05/2026", "end": "22/05/2026", "tags": "TR-06, FR-17",
     "epic": "E13: WorkDrive export", "branch": "feature/workdrive-prod-api (B4)",
     "desc": "AC: OAuth-authed upload of [Client]_[Provider]_[PlanRef]_[YYYYMMDD].xlsx; returns deep link stored on case; status auto-flips to Complete on success."},
    {"name": "RingCentral API → recording retrieval", "owner": "backendDev", "priority": "High", "status": "Open",
     "start": "15/05/2026", "end": "22/05/2026", "tags": "TR-05",
     "epic": "E14: RingCentral + Palindrome", "branch": "feature/ringcentral-palindrome (B5)",
     "desc": "AC: Fetch from RingCentral retrieves recording by call ID; audio uploaded to internal storage; manual paste fallback retained."},
    {"name": "Palindrome integration for transcription", "owner": "backendDev", "priority": "High", "status": "Open",
     "start": "15/05/2026", "end": "22/05/2026", "tags": "TR-05, TR-11",
     "epic": "E14: RingCentral + Palindrome", "branch": "feature/ringcentral-palindrome (B5)",
     "deps": "RingCentral API → recording retrieval",
     "desc": "AC: audio handed to Palindrome API; transcript stored on case + audit log; degrades to manual paste on failure."},
    {"name": "Origo URL routing per provider in Stage 1", "owner": "frontendDev", "priority": "Medium", "status": "Open",
     "start": "15/05/2026", "end": "22/05/2026", "tags": "TR-12, FR-23",
     "epic": "E15: Origo & LOA", "branch": "feature/origo-and-loa (B6)",
     "desc": "AC: Stage 1 surfaces 'Open in Origo' button when isOnOrigo=true; falls back to 'Send via Email' with mailto for non-Origo; wet-sig providers show 'Print LOA' + postal address."},
    {"name": "Updated LOA template + email subject lines", "owner": "backendDev", "priority": "High", "status": "Open",
     "start": "15/05/2026", "end": "22/05/2026", "tags": "FR-21",
     "epic": "E15: Origo & LOA", "branch": "feature/origo-and-loa (B6)",
     "desc": "AC: LOA template includes policy ref, scheme name, plan type, GDPR consent. Subject auto-fills 'LOA – [Client] – [Provider] – [Policy ref]'. Ceding-dept email pre-filled, main cc'd."},
    {"name": "Stage 9 paraplanner notification (Zoho task + in-app)", "owner": "backendDev", "priority": "High", "status": "Open",
     "start": "15/05/2026", "end": "22/05/2026", "tags": "FR-16",
     "epic": "E16: Stage 9 paraplanner notification", "branch": "feature/notifications-stage9 (B7)",
     "deps": "Zoho CRM blueprint webhook → auto-create case",
     "desc": "AC: 'Mark as Ready for Review' creates Zoho task with deep link; in-app banner sent; re-review loop triggers fresh notification + audit entry."},
    {"name": "Azure Key Vault wiring (env wiring)", "owner": "securityDev, devOpsDev", "priority": "High", "status": "Open",
     "start": "15/05/2026", "end": "22/05/2026", "tags": "TR-08",
     "epic": "E17: Secrets management", "branch": "feature/secrets-keyvault (B8)",
     "desc": "AC: all secrets (DB, Azure OpenAI key, Zoho refresh, RingCentral, Palindrome, JWT) moved out of .env into Key Vault. Backend reads via managed identity in prod, file-based in dev. Rotation policy documented."},
    {"name": "Backend swap aiExtraction.ts Anthropic → Azure OpenAI", "owner": "backendDev", "priority": "High", "status": "Open",
     "start": "15/05/2026", "end": "22/05/2026", "tags": "TR-11",
     "epic": "E18: Backend AI swap", "branch": "feature/azure-openai-swap (B2)",
     "deps": "AI set up (Sprint 1)",
     "desc": "AC: aiExtraction.ts calls Nishant's Azure OpenAI proxy, not Anthropic. Same I/O contract maintained. All extraction unit tests pass. Anthropic SDK removed from prod build."},
    {"name": "Backend swap aiCallAssist.ts Anthropic → Azure OpenAI", "owner": "backendDev", "priority": "High", "status": "Open",
     "start": "15/05/2026", "end": "22/05/2026", "tags": "TR-11",
     "epic": "E18: Backend AI swap", "branch": "feature/azure-openai-swap (B2)",
     "deps": "AI set up (Sprint 1)",
     "desc": "AC: call script + transcript analysis hit Azure OpenAI. Same response shape. Tests pass."},
    # AI track (Nishant)
    {"name": "[AI] Pension extraction prompts (53 fields)", "owner": "Nishant R", "priority": "High", "status": "Open",
     "start": "15/05/2026", "end": "22/05/2026", "tags": "FR-10, FR-12",
     "epic": "E19: AI prompts & accuracy",
     "desc": "All Pension checklist fields per requirements §8 covered. Conditional logic respected (e.g. with-profits sub-section)."},
    {"name": "[AI] ISA extraction prompts (33 fields)", "owner": "Nishant R", "priority": "High", "status": "Open",
     "start": "15/05/2026", "end": "22/05/2026", "tags": "FR-10, FR-12",
     "epic": "E19: AI prompts & accuracy",
     "desc": "All ISA checklist fields per requirements §8 covered."},
    {"name": "[AI] GIA extraction prompts (28 fields)", "owner": "Nishant R", "priority": "High", "status": "Open",
     "start": "15/05/2026", "end": "22/05/2026", "tags": "FR-10, FR-12",
     "epic": "E19: AI prompts & accuracy",
     "desc": "All GIA checklist fields per requirements §8 covered."},
    {"name": "[AI] Confidence scoring + page citation", "owner": "Nishant R", "priority": "High", "status": "Open",
     "start": "15/05/2026", "end": "22/05/2026", "tags": "FR-08, FR-11",
     "epic": "E19: AI prompts & accuracy",
     "desc": "AC: each field returns {value, confidence, page, evidenceRef}. Confidence one of HIGH / MEDIUM / LOW / MISSING."},
    {"name": "[AI] Multi-document conflict detection", "owner": "Nishant R", "priority": "High", "status": "Open",
     "start": "15/05/2026", "end": "22/05/2026", "tags": "FR-07",
     "epic": "E19: AI prompts & accuracy",
     "desc": "AC: when second doc contradicts first, field flagged CONFLICT; CA team prompted to choose value to keep."},
    {"name": "[AI] Call script generation prompts", "owner": "Nishant R", "priority": "High", "status": "Open",
     "start": "15/05/2026", "end": "22/05/2026", "tags": "FR-18",
     "epic": "E19: AI prompts & accuracy",
     "desc": "AC: AI-generated script from missing/low-confidence fields with provider phone, dept, questions per missing field."},
    {"name": "[AI] Transcript analysis prompts (Stage 7)", "owner": "Nishant R", "priority": "High", "status": "Open",
     "start": "15/05/2026", "end": "22/05/2026", "tags": "FR-20",
     "epic": "E19: AI prompts & accuracy",
     "desc": "AC: AI matches answers to MISSING/LOW fields, pre-fills, tags each with call source (date + CA user name)."},
    {"name": "Integration test fixtures (Zoho/RingCentral/WorkDrive mocks)", "owner": "testerDev", "priority": "Medium", "status": "Open",
     "start": "15/05/2026", "end": "22/05/2026", "tags": "NFR-01",
     "epic": "E19: AI prompts & accuracy",
     "desc": "AC: MSW or nock-based mocks; tests run hermetic (no network)."},
]
for n, t in enumerate(S2, start=1):
    t["num"] = n
add_sprint_sheet("Sprint 2 Items", S2)

# ───────────── Sprint 3 (23 May – 1 June) ─────────────
S3 = [
    # Azure infra
    {"name": "Azure App Service + Static Web Apps", "owner": "devOpsDev", "priority": "High", "status": "Open",
     "start": "23/05/2026", "end": "01/06/2026", "tags": "TR-01",
     "epic": "E20: Azure infrastructure", "branch": "feature/azure-infrastructure (B9)",
     "desc": "AC: backend on App Service Linux Node 20; frontend on Static Web Apps with custom domain; both wired to App Insights."},
    {"name": "Azure Database for PostgreSQL (Flexible Server)", "owner": "devOpsDev, databaseDev", "priority": "High", "status": "Open",
     "start": "23/05/2026", "end": "01/06/2026", "tags": "TR-01, TR-08",
     "epic": "E20: Azure infrastructure", "branch": "feature/azure-infrastructure (B9)",
     "desc": "AC: Postgres 15+ Flexible Server in Furnley tenant; private endpoint, firewall locked to App Service; automated backups."},
    {"name": "Azure Blob Storage for documents", "owner": "devOpsDev", "priority": "High", "status": "Open",
     "start": "23/05/2026", "end": "01/06/2026", "tags": "TR-01, TR-08",
     "epic": "E20: Azure infrastructure", "branch": "feature/azure-infrastructure (B9)",
     "desc": "AC: blob container with AES-256 server-side encryption; SAS-based access from backend."},
    {"name": "Azure OpenAI deployment + private endpoint", "owner": "devOpsDev, Nishant R", "priority": "High", "status": "Open",
     "start": "23/05/2026", "end": "01/06/2026", "tags": "TR-01, TR-11",
     "epic": "E20: Azure infrastructure", "branch": "feature/azure-infrastructure (B9)",
     "desc": "AC: Azure OpenAI provisioned in Furnley tenant; GPT-4o (or equivalent) deployed; private endpoint."},
    {"name": "Application Insights + alerts", "owner": "devOpsDev", "priority": "High", "status": "Open",
     "start": "23/05/2026", "end": "01/06/2026", "tags": "NFR-03",
     "epic": "E20: Azure infrastructure", "branch": "feature/azure-infrastructure (B9)",
     "desc": "AC: backend + frontend traces, requests, exceptions; alerts: error >1%, p95 >5s, availability <99.5%; SLT-visible dashboard."},
    {"name": "TLS cert + custom domain (ceding.furnleyhouse.co.uk)", "owner": "devOpsDev, securityDev", "priority": "High", "status": "Open",
     "start": "23/05/2026", "end": "01/06/2026", "tags": "TR-08",
     "epic": "E20: Azure infrastructure", "branch": "feature/azure-infrastructure (B9)",
     "desc": "AC: TLS 1.2+ enforced; HSTS header set; auto-renewal via Azure-managed cert."},
    {"name": "Auto-scaling rules + budget alerts", "owner": "devOpsDev", "priority": "Medium", "status": "Open",
     "start": "23/05/2026", "end": "01/06/2026", "tags": "NFR-06",
     "epic": "E20: Azure infrastructure", "branch": "feature/azure-infrastructure (B9)",
     "desc": "AC: scale-out at CPU >70% (max 4 instances during business hours); budget alert at 50/80/100% of monthly cap."},
    {"name": "Postgres backup / restore procedure", "owner": "devOpsDev, databaseDev", "priority": "High", "status": "Open",
     "start": "23/05/2026", "end": "01/06/2026", "tags": "TR-08",
     "epic": "E20: Azure infrastructure", "branch": "feature/azure-infrastructure (B9)",
     "desc": "AC: daily automated backup, 35-day retention; restore drill on staging documented in runbook."},
    # CI/CD
    {"name": "GitHub Actions CI/CD (build/test/deploy)", "owner": "devOpsDev", "priority": "High", "status": "Open",
     "start": "23/05/2026", "end": "01/06/2026", "tags": "TR-01",
     "epic": "E21: CI/CD pipeline", "branch": "feature/ci-cd-pipeline (B10)",
     "deps": "Azure App Service + Static Web Apps",
     "desc": "AC: PR pipeline lint+typecheck+unit; merge→dev deploy; release/* tag → staging; v*.*.* tag → prod with manual approval gate."},
    {"name": "Environments dev/staging/prod separate config", "owner": "devOpsDev", "priority": "High", "status": "Open",
     "start": "23/05/2026", "end": "01/06/2026", "tags": "TR-01",
     "epic": "E21: CI/CD pipeline", "branch": "feature/ci-cd-pipeline (B10)",
     "desc": "AC: 3 resource groups, 3 Key Vaults, 3 Postgres servers; per-env DNS."},
    # Retention
    {"name": "1-year automated retention deletion job", "owner": "backendDev, devOpsDev", "priority": "High", "status": "Open",
     "start": "23/05/2026", "end": "01/06/2026", "tags": "TR-10",
     "epic": "E22: Data retention", "branch": "feature/data-retention (B11)",
     "desc": "AC: daily Azure Function deletes documents + AI artefacts >365 days old; audit log entry per deletion; manual override; aligned to FCA COBS."},
    # Provider data
    {"name": "Production DB migration + seed (providers, templates, users)", "owner": "databaseDev, backendDev", "priority": "High", "status": "Open",
     "start": "23/05/2026", "end": "01/06/2026", "tags": "TR-01",
     "epic": "E23: Provider directory data", "branch": "feature/provider-directory-load (B12)",
     "desc": "AC: all Prisma migrations applied to staging+prod; default checklist templates seeded for Pension/ISA/GIA per §8; initial admin user created."},
    {"name": "Provider Directory full data load", "owner": "backendDev", "priority": "High", "status": "Open",
     "start": "23/05/2026", "end": "01/06/2026", "tags": "FR-19, Q-06",
     "epic": "E23: Provider directory data", "branch": "feature/provider-directory-load (B12)",
     "deps": "Aruna delivers verified provider file",
     "desc": "AC: all Origo providers loaded (§9.2); known contacts loaded (§9.3); LOA signature categorisation applied (§9.1); verified by Aruna pre-launch."},
    {"name": "Policy reference prefix routing rules", "owner": "backendDev", "priority": "High", "status": "Open",
     "start": "23/05/2026", "end": "01/06/2026", "tags": "FR-19, Q-07",
     "epic": "E23: Provider directory data", "branch": "feature/provider-directory-load (B12)",
     "deps": "Aruna confirms prefix patterns",
     "desc": "AC: planTypePrefixes[] populated per provider; Stage 1 routes correctly based on prefix match; audit log records routing decisions."},
    # UI polish
    {"name": "Final UI polish — Inter font, navy/teal branding", "owner": "frontendDev", "priority": "Medium", "status": "Open",
     "start": "23/05/2026", "end": "01/06/2026", "tags": "UI/UX",
     "epic": "E24: UI launch polish", "branch": "feature/ui-launch-polish (B13)",
     "desc": "AC: Inter loaded; primary #0D1B2A; accent #00C2CB; Furnley lion logo in navbar; consistent card-based layout."},
    {"name": "Empty states + skeleton loaders + global error toast", "owner": "frontendDev", "priority": "Medium", "status": "Open",
     "start": "23/05/2026", "end": "01/06/2026", "tags": "UI/UX",
     "epic": "E24: UI launch polish", "branch": "feature/ui-launch-polish (B13)",
     "desc": "AC: every list page has skeleton + empty state with CTA; AI extraction failure shows red banner with retry; required-field errors highlighted in red on save."},
    {"name": "Accessibility audit (WCAG AA basics)", "owner": "testerDev, frontendDev", "priority": "Medium", "status": "Open",
     "start": "23/05/2026", "end": "01/06/2026", "tags": "UI/UX",
     "epic": "E24: UI launch polish", "branch": "feature/ui-launch-polish (B13)",
     "desc": "AC: Axe scan passes on all main pages; keyboard nav works for case detail + admin panels; colour contrast ≥4.5:1 for text."},
    # Security
    {"name": "Security review (TLS, headers, secrets, RBAC matrix)", "owner": "securityDev", "priority": "High", "status": "Open",
     "start": "23/05/2026", "end": "01/06/2026", "tags": "TR-07, TR-08",
     "epic": "E25: Security & GDPR",
     "desc": "AC: TLS 1.2+ enforced; HSTS/CSP/X-Frame-Options set; no secrets in code or env; RBAC matrix matches §6 line by line; high-severity findings blocked from prod."},
    {"name": "GDPR sign-off pack — handed to John/Lee", "owner": "securityDev", "priority": "High", "status": "Open",
     "start": "23/05/2026", "end": "01/06/2026", "tags": "TR-09",
     "epic": "E25: Security & GDPR",
     "desc": "HARD BLOCKER for prod launch. AC: DPIA completed; data flow diagram; retention policy aligned to FCA COBS; submitted to John Hood / Lee Parkinson."},
    # AI accuracy gate
    {"name": "AI accuracy benchmark across 50 PDFs / 6+ providers", "owner": "Nishant R, testerDev", "priority": "High", "status": "Open",
     "start": "23/05/2026", "end": "01/06/2026", "tags": "NFR-04",
     "epic": "E26: AI accuracy gate",
     "deps": "AI training (Sprint 2)",
     "desc": "HARD GATE for UAT. AC: 50+ representative PDFs from 6+ providers staged; per-field confidence scored; ≥85% HIGH-confidence rate per field type."},
    {"name": "Prompt tuning iterations to reach 85% HIGH confidence", "owner": "Nishant R", "priority": "High", "status": "Open",
     "start": "23/05/2026", "end": "01/06/2026", "tags": "NFR-04",
     "epic": "E26: AI accuracy gate",
     "desc": "Continues from Sprint 2 AI training. Iterate prompts until benchmark gate passes."},
    # Performance
    {"name": "E2E tests for full 10-stage flow (Playwright)", "owner": "testerDev", "priority": "High", "status": "Open",
     "start": "23/05/2026", "end": "01/06/2026", "tags": "NFR-01",
     "epic": "E27: Performance & load", "branch": "feature/perf-tuning (B14)",
     "desc": "AC: suite walks Case Create → Stage 1 → ... → Stage 10 export; runs in CI on every PR; records video on failure."},
    {"name": "Performance benchmarks (page load <3s, AI <2min)", "owner": "testerDev", "priority": "High", "status": "Open",
     "start": "23/05/2026", "end": "01/06/2026", "tags": "NFR-01, NFR-02",
     "epic": "E27: Performance & load", "branch": "feature/perf-tuning (B14)",
     "desc": "AC: page load p95 <3s on 10 Mbps (UK + Chennai); AI extraction p95 <2 min for 20-page PDFs; results recorded in runbook."},
    {"name": "Load test for 20 concurrent users", "owner": "testerDev", "priority": "High", "status": "Open",
     "start": "23/05/2026", "end": "01/06/2026", "tags": "NFR-06",
     "epic": "E27: Performance & load", "branch": "feature/perf-tuning (B14)",
     "desc": "AC: k6/Artillery script simulates 20 concurrent CA users; backend p95 <1s under load; no 5xx; auto-scale rules trigger."},
    # Cutover prep
    {"name": "Production deployment runbook + rollback plan", "owner": "devOpsDev", "priority": "High", "status": "Open",
     "start": "23/05/2026", "end": "01/06/2026", "tags": "TR-01",
     "epic": "E28: Cutover prep & code freeze",
     "desc": "AC: step-by-step cutover; rollback procedure (release tag revert + DB restore); DNS cutover plan; on-call rota documented."},
    {"name": "IT test (dev integration) — Zoho task", "owner": "Revathy S", "priority": "High", "status": "Open",
     "start": "22/05/2026", "end": "28/05/2026", "tags": "—",
     "epic": "E28: Cutover prep & code freeze", "branch": "release/v1.0-rc (B15)",
     "desc": "Zoho task (AT RISK). AC: dev team validates end-to-end happy path; bugs raised + triaged before code freeze; sign-off before 1 June."},
    {"name": "Code freeze 1 June — RC tag, handover to internal testing", "owner": "All", "priority": "High", "status": "Open",
     "start": "01/06/2026", "end": "01/06/2026", "tags": "—",
     "epic": "E28: Cutover prep & code freeze", "branch": "release/v1.0-rc (B15)",
     "desc": "AC: all S1–S3 branches merged; release/v1.0-rc tag pushed; backend+frontend on staging; hand-off doc to Chennai CA team."},
]
for n, t in enumerate(S3, start=1):
    t["num"] = n
add_sprint_sheet("Sprint 3 Items", S3)

# ───────────── Sprint 4 (1–7 June) ─────────────
S4 = [
    {"name": "AI accuracy gate verification (final, 85% HIGH)", "owner": "Nishant R, testerDev", "priority": "High", "status": "Open",
     "start": "01/06/2026", "end": "07/06/2026", "tags": "NFR-04",
     "epic": "E29: AI accuracy gate (final)",
     "desc": "HARD GATE for UAT. AC: final benchmark on staged PDFs; ≥85% HIGH per field; sign-off from Lee + Aruna."},
    {"name": "CA functional UAT plan + scenario script", "owner": "testerDev", "priority": "High", "status": "Open",
     "start": "01/06/2026", "end": "07/06/2026", "tags": "NFR-05",
     "epic": "E30: CA team functional UAT",
     "desc": "AC: test cases cover all 10 stages; one full test per plan type (Pension/ISA/GIA); edge cases (re-review, conflict, On Hold)."},
    {"name": "Test data pack — 50 PDFs from 6+ providers staged", "owner": "testerDev", "priority": "High", "status": "Open",
     "start": "01/06/2026", "end": "07/06/2026", "tags": "NFR-04",
     "epic": "E30: CA team functional UAT",
     "desc": "Representative PDFs staged in fixtures; sufficient for benchmark + UAT."},
    {"name": "Daily bug triage + fix rota (BE/FE/AI/DevOps rotation)", "owner": "All", "priority": "High", "status": "Open",
     "start": "01/06/2026", "end": "07/06/2026", "tags": "—",
     "epic": "E30: CA team functional UAT", "branch": "hotfix/* (B16)",
     "desc": "AC: daily 30-min standup; bug board prioritised; blockers <24h fix SLA; all fixes covered by tests."},
    {"name": "Daily test status report to Lee + Aruna", "owner": "testerDev", "priority": "High", "status": "Open",
     "start": "01/06/2026", "end": "07/06/2026", "tags": "—",
     "epic": "E30: CA team functional UAT",
     "desc": "Daily 1-page report: pass/fail counts, blockers, ETA on red items."},
    {"name": "Live security smoke test on staging", "owner": "securityDev", "priority": "High", "status": "Open",
     "start": "01/06/2026", "end": "07/06/2026", "tags": "TR-07, TR-08",
     "epic": "E30: CA team functional UAT",
     "desc": "AC: OWASP Top 10 quick checks; auth bypass attempts; SQL injection / XSS spot checks."},
    {"name": "IT bug fix (long-running, S3–S6)", "owner": "Nishant R, Srinath K, Revathy S", "priority": "High", "status": "Open",
     "start": "20/05/2026", "end": "03/07/2026", "tags": "—",
     "epic": "E31: Long-running IT bug fix",
     "desc": "Zoho task. Daily bug board kept current; all Sev 1/2 fixed before next gate; Sev 3/4 carried as Phase 2 candidates."},
]
for n, t in enumerate(S4, start=1):
    t["num"] = n
add_sprint_sheet("Sprint 4 Items", S4)

# ───────────── Sprint 5 (8–15 June) ─────────────
S5 = [
    {"name": "UAT scenarios authored with business sign-off owners", "owner": "testerDev", "priority": "High", "status": "Open",
     "start": "08/06/2026", "end": "15/06/2026", "tags": "NFR-05",
     "epic": "E32: Business UAT",
     "desc": "AC: per-stakeholder scenarios (Lee, Natalie, Aruna, Matt, Nicki); covers paraplanner approval flow, adviser handoff, export."},
    {"name": "UAT facilitation (5 business users)", "owner": "testerDev", "priority": "High", "status": "Open",
     "start": "08/06/2026", "end": "15/06/2026", "tags": "NFR-05",
     "epic": "E32: Business UAT",
     "desc": "Facilitate scenario walkthroughs; record outcomes; capture bugs."},
    {"name": "UAT bug triage + fix", "owner": "All", "priority": "High", "status": "Open",
     "start": "08/06/2026", "end": "15/06/2026", "tags": "—",
     "epic": "E32: Business UAT", "branch": "hotfix/* (B18)",
     "desc": "Same daily rota as Sprint 4."},
    {"name": "UAT sign-off package to Lee Parkinson", "owner": "testerDev", "priority": "High", "status": "Open",
     "start": "08/06/2026", "end": "15/06/2026", "tags": "NFR-05",
     "epic": "E32: Business UAT",
     "desc": "AC: signed-off scenario list; open-bug register with target fix dates; Lee's go/no-go recorded."},
    {"name": "Final pen-test on staging", "owner": "securityDev", "priority": "High", "status": "Open",
     "start": "08/06/2026", "end": "15/06/2026", "tags": "TR-08",
     "epic": "E33: Final security",
     "desc": "AC: external pen-tester (or internal sec); findings documented; high-severity issues fixed before cutover."},
    {"name": "Production cutover plan + rollback drill", "owner": "devOpsDev", "priority": "High", "status": "Open",
     "start": "08/06/2026", "end": "15/06/2026", "tags": "TR-01",
     "epic": "E34: Cutover preparation",
     "desc": "AC: step-by-step cutover script (DNS, secrets, data migration); rollback drill on staging; cutover window agreed with Lee + Natalie."},
    {"name": "Final data migration scripts (provider directory, templates)", "owner": "databaseDev", "priority": "High", "status": "Open",
     "start": "08/06/2026", "end": "15/06/2026", "tags": "TR-01",
     "epic": "E34: Cutover preparation",
     "desc": "AC: idempotent migration scripts; validation queries to confirm prod state matches expected."},
]
for n, t in enumerate(S5, start=1):
    t["num"] = n
add_sprint_sheet("Sprint 5 Items", S5)

# ───────────── Sprint 6 (16–27 June) ─────────────
S6 = [
    {"name": "Production cutover (16 or 17 June)", "owner": "devOpsDev, All", "priority": "High", "status": "Open",
     "start": "16/06/2026", "end": "17/06/2026", "tags": "TR-01",
     "epic": "E35: Cutover", "branch": "release/v1.0 (B19) tagged v1.0.0",
     "deps": "UAT sign-off",
     "desc": "Zoho Projects task currently shows 10 Jun — UPDATE to 16/17 Jun. AC: cutover script executed; smoke tests pass on prod; DNS active; rollback ready."},
    {"name": "DNS + traffic routing", "owner": "devOpsDev", "priority": "High", "status": "Open",
     "start": "16/06/2026", "end": "17/06/2026", "tags": "TR-08",
     "epic": "E35: Cutover",
     "desc": "Cutover DNS to prod environment; verify TTL; pre-warm CDN if applicable."},
    {"name": "Production smoke tests post-deploy", "owner": "testerDev", "priority": "High", "status": "Open",
     "start": "16/06/2026", "end": "17/06/2026", "tags": "NFR-03",
     "epic": "E35: Cutover",
     "desc": "AC: login + create case + extract + export end-to-end; smoke pack runs in <15 min; all green before SLT announcement."},
    {"name": "Post-deploy security audit", "owner": "securityDev", "priority": "High", "status": "Open",
     "start": "16/06/2026", "end": "17/06/2026", "tags": "TR-08",
     "epic": "E35: Cutover",
     "desc": "Quick security audit on prod environment after cutover."},
    {"name": "Hypercare daily standup (10 working days)", "owner": "All", "priority": "High", "status": "Open",
     "start": "18/06/2026", "end": "27/06/2026", "tags": "—",
     "epic": "E36: Hypercare",
     "desc": "Daily 30-min standup; review issues, monitoring, hot-fixes."},
    {"name": "On-call rota (BE/FE/AI/DevOps)", "owner": "All", "priority": "High", "status": "Open",
     "start": "18/06/2026", "end": "27/06/2026", "tags": "—",
     "epic": "E36: Hypercare",
     "desc": "AC: 24/5 cover during UK business hours; pager rotation defined and shared."},
    {"name": "Hot-fix patches (BE) as needed", "owner": "backendDev", "priority": "High", "status": "Open",
     "start": "18/06/2026", "end": "27/06/2026", "tags": "—",
     "epic": "E36: Hypercare", "branch": "hotfix/hypercare-* (B20)",
     "desc": "Double-merged to release/v1.0 + main."},
    {"name": "Hot-fix patches (FE) as needed", "owner": "frontendDev", "priority": "High", "status": "Open",
     "start": "18/06/2026", "end": "27/06/2026", "tags": "—",
     "epic": "E36: Hypercare", "branch": "hotfix/hypercare-* (B20)",
     "desc": "Double-merged to release/v1.0 + main."},
    {"name": "Production monitoring + alerts active (App Insights)", "owner": "devOpsDev", "priority": "High", "status": "Open",
     "start": "18/06/2026", "end": "27/06/2026", "tags": "NFR-03",
     "epic": "E36: Hypercare",
     "desc": "AC: 99.5% uptime hit during UK business hours; Sev 1 incidents resolved <4h; metrics summarised daily."},
    {"name": "KPI dashboard for SLT (cases/week, AI accuracy, calls/case)", "owner": "backendDev", "priority": "High", "status": "Open",
     "start": "18/06/2026", "end": "27/06/2026", "tags": "BR-06",
     "epic": "E37: KPI dashboard",
     "desc": "AC: live dashboard accessible to SLT; tracks cases/week vs 78 target; AI HIGH-conf vs 85%; calls per case vs 1."},
]
for n, t in enumerate(S6, start=1):
    t["num"] = n
add_sprint_sheet("Sprint 6 Items", S6)

# ───────────── Sprint 7 (28 June – 1 July) ─────────────
S7 = [
    {"name": "KPI metrics extraction (cases/week, time saved, AI accuracy, calls/case)", "owner": "backendDev, testerDev", "priority": "High", "status": "Open",
     "start": "28/06/2026", "end": "01/07/2026", "tags": "BR-05, BR-06",
     "epic": "E38: Success metrics & closeout",
     "desc": "AC: hard numbers vs baselines for all 4 KPIs; comparison prototype vs production; report shared with Lee + Natalie pre-meeting."},
    {"name": "Lessons learned doc compiled", "owner": "All", "priority": "Medium", "status": "Open",
     "start": "28/06/2026", "end": "01/07/2026", "tags": "—",
     "epic": "E38: Success metrics & closeout",
     "desc": "AC: per-track retro (FE / BE / DB / AI / DevOps / Sec / QA); 'what went well / didn't / would change'."},
    {"name": "Present success at June team event (Zoho task — 1 July)", "owner": "Revathy S", "priority": "High", "status": "Open",
     "start": "01/07/2026", "end": "01/07/2026", "tags": "—",
     "epic": "E38: Success metrics & closeout",
     "desc": "AC: slide deck prepared; KPI numbers + UAT outcomes shown; Phase 2 roadmap previewed."},
    {"name": "Phase 2 backlog grooming", "owner": "backendDev, All", "priority": "Medium", "status": "Open",
     "start": "28/06/2026", "end": "01/07/2026", "tags": "—",
     "epic": "E39: Phase 2 grooming",
     "desc": "AC: Bond, Final Salary/DB, Protection epics drafted; EIS/VCT/SR auto as separate workstream; sizing for next initiation."},
    {"name": "Hypercare metrics + post-launch retro report", "owner": "devOpsDev", "priority": "Medium", "status": "Open",
     "start": "28/06/2026", "end": "01/07/2026", "tags": "NFR-03",
     "epic": "E38: Success metrics & closeout",
     "desc": "Hypercare KPIs (uptime, MTTR, hot-fix count) + retro themes."},
    {"name": "Phase 2 security plan", "owner": "securityDev", "priority": "Medium", "status": "Open",
     "start": "28/06/2026", "end": "01/07/2026", "tags": "—",
     "epic": "E39: Phase 2 grooming",
     "desc": "Security plan for next phase: pen-test cadence, secrets rotation, audit posture."},
]
for n, t in enumerate(S7, start=1):
    t["num"] = n
add_sprint_sheet("Sprint 7 Items", S7)

# ===========================================================================
# Sheet — Branch Plan
# ===========================================================================
ws = wb.create_sheet("Branch Plan")
ws.sheet_view.showGridLines = False
title(ws, 1, "Branch Plan — B1 → B20")
note(ws, 2, "Branches off main, with release/* and hotfix/* branches for testing and launch windows.")
header(ws, 4, ["#", "Branch", "Sprint", "Purpose", "Merge after"], widths=[5, 38, 8, 70, 36])
branches = [
    ("B1", "feature/foundation-hardening", 1, "Zod audit, indexes, test scaffold + CI gate, dev-workflow docs", "—"),
    ("B2", "feature/azure-openai-swap", 2, "Backend swap Anthropic → Azure OpenAI", "B1; depends on Nishant's AI endpoint"),
    ("B3", "feature/zoho-blueprint-webhook", 2, "Auto-create case via webhook", "B1"),
    ("B4", "feature/workdrive-prod-api", 2, "WorkDrive live wiring", "B1"),
    ("B5", "feature/ringcentral-palindrome", 2, "RingCentral retrieval + Palindrome transcription", "B1; depends on Palindrome wrapper"),
    ("B6", "feature/origo-and-loa", 2, "Origo URL routing + LOA template + email subjects", "B1"),
    ("B7", "feature/notifications-stage9", 2, "Stage 9 paraplanner notification", "B3"),
    ("B8", "feature/secrets-keyvault", 2, "Azure Key Vault wiring", "B1"),
    ("B9", "feature/azure-infrastructure", 3, "App Service, Static Web Apps, Postgres, Blob, Insights, TLS, scaling, backups", "B8"),
    ("B10", "feature/ci-cd-pipeline", 3, "GitHub Actions + dev/staging/prod env config", "B9"),
    ("B11", "feature/data-retention", 3, "1-year retention deletion job", "B9"),
    ("B12", "feature/provider-directory-load", 3, "Prod migration + seed + provider data load + prefix routing", "B1"),
    ("B13", "feature/ui-launch-polish", 3, "Branding, empty states, error toast, a11y audit", "B2–B7"),
    ("B14", "feature/perf-tuning", 3, "E2E + perf benchmarks + load tests", "All Sprint 2 branches + B13"),
    ("B15", "release/v1.0-rc", 3, "RC cut at code freeze (1 Jun)", "After B1–B14"),
    ("B16", "hotfix/*", 4, "Bugs from Chennai CA testing — cherry-pick to main, merge forward", "—"),
    ("B17", "release/v1.0-rc2", 5, "UAT-fix cut from B15 after Chennai sign-off", "B15 + Sprint 4 hotfixes"),
    ("B18", "hotfix/*", 5, "Bugs from business UAT", "—"),
    ("B19", "release/v1.0", 6, "Production cutover, tagged v1.0.0", "B17 + Sprint 5 hotfixes"),
    ("B20", "hotfix/hypercare-*", 6, "10-day hypercare patches; double-merged release/v1.0 + main", "—"),
]
for i, b in enumerate(branches, start=5):
    row(ws, i, list(b))
    ws.row_dimensions[i].height = 32

# ── Save ───────────────────────────────────────────────────
wb.save(OUT)
print(f"Saved: {OUT}")
print(f"Sheets: {wb.sheetnames}")
total_items = len(S1) + len(S2) + len(S3) + len(S4) + len(S5) + len(S6) + len(S7)
print(f"Total items across sprints: {total_items}")
print(f"Epics: {len(epics)}")
print(f"Branches: {len(branches)}")
